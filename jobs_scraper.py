from addon import *
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from openpyxl import Workbook

Titles=[]
Links=[]
Locations=[]
Details=[]
# CHƯA XONG, CHƯA XONG, CHƯA XONG, CHƯA XONG, CHƯA XONG
def get_job_data( value):
    headless=True
    driver = init_browser(headless=headless)

    url = 'https://www.careerbuilder.com/jobs-ceo-in-los-angeles,ca'
    driver.get(url)
    # Lưu trữ chỉ số của phần tử cuối cùng trong danh sách
    last_index = -1

    while True:
        if len(Links)>value-1 or value==0:
                break
        print(f'Load  lấy 25 job')
        # Chờ cho danh sách công việc được tải đầy đủ vào "Job_elements"
        job_elements = WebDriverWait(driver, 20).until(EC.presence_of_all_elements_located((By.CLASS_NAME, "data-results-content-parent")),"ace")

        # Lặp qua các phần tử công việc để lấy thông tin chi tiết của các công việc mới
        for j in range(last_index+1, len(job_elements)):
            job = job_elements[j]
            job.click()
            time.sleep(1)
            # Lấy thông tin công việc 
            
            link=job.find_element(By.TAG_NAME,'a').get_attribute('href')
            title='not found'
            location='not found'
            detail='not found'
            try:
                title=job.find_element(By.XPATH,'//*[@id="jdp-data"]/div[1]/div[2]/div/div[1]/h2').text
            except:
                pass
            try:
                detail=driver.find_element(By.CSS_SELECTOR,'#jdp_description').text
            except:
                pass
            try:
                location=job.find_element(By.XPATH,'//*[@id="jdp-data"]/div[1]/div[2]/div/div[1]/div[1]/span[2]').text
            except:
                pass
            
            Titles.append(title)
            Links.append(link)
            Locations.append(location)
            Details.append(detail)

            if len(Links)>value-1:
                break
        try:
            load_more_button =driver.find_element(By.CSS_SELECTOR,'#load_more_jobs')
            if not load_more_button.is_displayed():
                break
        except:
            break
        # Lấy chỉ số của phần tử cuối cùng trong danh sách các job trước đó
        last_job_element = job_elements[last_index]
        last_index = job_elements.index(last_job_element)

        load_more_button.click()

        time.sleep(3)
        
    output_Excel()
    time.sleep(2)
    # Đóng trình duyệt
    driver.quit()
    
def output_Excel():
    print("tiến hành lưu")
    wb = Workbook()

    sheet = wb.active
    #Tên cột
    sheet['A1'] = 'num order'
    sheet['B1'] = 'Link'
    sheet['C1'] = 'Title'
    sheet['D1'] = 'location'
    sheet['E1'] = 'detail'
    for i in range(0, len(Links)):  # cột
        sheet.cell(column=1, row=i + 2, value=i+1)
    #lưu cột chứa link
    for i in range(0, len(Links)):  # cột
        v = Links[i]
        sheet.cell(column=2, row=i + 2, value=v)
    #lưu cột chứa title
    for i in range(0, len(Links)):  # cột
        v = Titles[i]
        sheet.cell(column=3, row=i + 2, value=v)
    #lưu cột chứa location
    for i in range(0, len(Links)):  # cột
        v = Locations[i]
        sheet.cell(column=4, row=i + 2, value=v)
    #lưu cột chứa detail
    for i in range(0, len(Links)):  # cột
        v = Details[i]
        sheet.cell(column=5, row=i + 2, value=v)
     # Lấy đường dẫn thư mục chứa chương trình Python
    program_dir = os.path.dirname(os.path.abspath(__file__))
 
    folder_path=os.path.join(program_dir,"..","..","..","output")
    file_name="jobs1.xlsx"
    # Tạo đường dẫn đầy đủ đến file mới
    file_path = os.path.join(folder_path, file_name)
    
    wb.save(file_path)
    print("lưu thành công")
    print("File đã được lưu tại:", file_path)




